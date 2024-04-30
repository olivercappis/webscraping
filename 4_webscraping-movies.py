
from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font





#webpage = 'https://www.boxofficemojo.com/weekend/chart/'
webpage = 'https://www.boxofficemojo.com/year/2024/'

page = urlopen(webpage)			

soup = BeautifulSoup(page, 'html.parser')

title = soup.title

print(title.text)
##
##
##
##
table_rows = soup.findAll('tr')

wb = xl.Workbook()
ws = wb.active
ws.title = 'Box Office Report'
ws['A1'] = 'No.'
ws['B1'] = 'Movie Title'
ws['C1'] = 'Release Date'
ws['D1'] = 'Number of Theaters'
ws['E1'] = 'Total Gross'
ws['F1'] = 'Average Gross by Theater'

for row in range(1,6):
    td = table_rows[row].findAll('td')
    number = td[0].text
    title = td[1].text
    release_date = td[8].text
    number_of_theater = int(td[6].text.replace(',',''))
    total_gross = int(td[7].text.replace(',','').replace('$',''))
    average_gross_theater = total_gross / number_of_theater

    ws['A' + str(row + 1)] = number
    ws['B' + str(row + 1)] = title
    ws['C' + str(row + 1)] = release_date
    ws['D' + str(row + 1)] = number_of_theater
    ws['E' + str(row + 1)] = total_gross
    ws['F' + str(row + 1)] = average_gross_theater


header_font = Font(size = 16, bold = True)
for cell in ws[1:1]:
    cell.font = header_font