import openpyxl as xl
from openpyxl.styles import Font

#create new document
wb = xl.Workbook()

ws = wb.active

ws.title = 'First Sheet'

wb.create_sheet(index=1, title= 'Second Sheet')

#write content to a cell
ws['A1'] = 'Invoice'

ws['A1'].font = Font(name= 'Times New Roman', size = 24, bold = True)

headerfont = Font(name= 'Times New Roman', size = 24, bold = True)

ws['A2'] = 'Tires'
ws['A3'] = 'Brakes'
ws['A4'] = "Alignment"

ws.merge_cells('A1:B1')

ws['B2'] = 450
ws['B3'] = 225
ws['B4'] = 150

#unmerge cells
#ws.unmerge_cells("A1:B1")

ws['A8'] = 'Total'
ws['A8'].font = Font(size = 16,bold=True)

ws.column_dimensions['A'].width = 25

ws['B8'] = '=SUM(B2:B4)'








write_sheet = wb['Second Sheet']
read_wb = xl.load_workbook("ProduceReport.xlsx")
reader = read_wb.active
read_ws = read_wb['ProduceReport']

for row in reader.iter_rows():
    values = [cell.value for cell in row]
    write_sheet.append(values)

wb['Second Sheet']['A43'] = 'Totals'
wb['Second Sheet']['C43'] = '=SUM(C2:C41)'
wb['Second Sheet']['D43'] = '=SUM(D2:D41)'

wb['Second Sheet']['A44'] = 'Averages'
wb['Second Sheet']['C44'] = '=AVERAGE(C2:C41)'
wb['Second Sheet']['D44'] = '=AVERAGE(D2:D41)'





wb.save("pythontoexcel.xlsx")